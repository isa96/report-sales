import pandas as pd
from openpyxl import Workbook
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
import string


def automate_excel(file_name, sheet_name, index, chart_style, columns=None):
    """The file name should have the following structure: sales_month.xlsx"""
    # read excel file
    excel_file = pd.read_excel(file_name)
    # make pivot table
    report_table = excel_file.pivot_table(index=index, columns=columns, values='Total', aggfunc='sum').round(0)
    # splitting the month and extension from the file name
    month_and_extension = file_name.split('_')[1]

    # send the report table to excel file
    try:
        with pd.ExcelWriter(f'output/report_{month_and_extension}', mode='a', if_sheet_exists='replace') as writer:
            report_table.to_excel(writer, sheet_name=sheet_name, startrow=4)
    except KeyError:
        with pd.ExcelWriter(f'output/report_{month_and_extension}') as writer:
            report_table.to_excel(writer, sheet_name=sheet_name, startrow=4)
    # loading workbook and selecting sheet
    wb = load_workbook(f'output/report_{month_and_extension}')
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.active = wb[sheet_name]
    sheet = wb[sheet_name]
    # cell references (original spreadsheet)
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    # adding a chart
    chart_style = chart_style #BarChart()
    data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row) #including headers
    categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row) #not including headers
    chart_style.add_data(data, titles_from_data=True)
    chart_style.set_categories(categories)
    sheet.add_chart(chart_style, "B12") #location chart
    chart_style.title = 'Sales by {}'.format(sheet_name)
    chart_style.style = 2 #choose the chart style
    # applying formulas
    # first create alphabet list as references for cells
    alphabet = list(string.ascii_uppercase)
    excel_alphabet = alphabet[0:max_column] #note: Python lists start on 0 -> A=0, B=1, C=2. #note2 the [a:b] takes b-a elements
    # sum in columns B-G
    for i in excel_alphabet:
        if i!='A':
            sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
            sheet[f'{i}{max_row+1}'].style = 'Currency'
    sheet[f'{excel_alphabet[0]}{max_row+1}'] = 'Total'
    # getting month name
    month_name = month_and_extension.split('.')[0]
    # formatting the report
    sheet['A1'] = 'Sales Report'
    sheet['A2'] = month_name.title()
    sheet['A1'].font = Font('Arial', bold=True, size=20)
    sheet['A2'].font = Font('Arial', bold=True, size=10)
    wb.save(f'output/report_{month_and_extension}')



automate_excel('data/sales_2021.xlsx', 'Product line', 'Gender', BarChart(), 'Product line')
automate_excel('data/sales_2021.xlsx', 'City', 'City', PieChart())